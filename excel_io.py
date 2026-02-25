from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEET_NAME = "Deptos"


DEFAULT_COLUMNS = [
    # identity
    "id",
    "nombre",
    "zona_colonia",
    "fuente",
    "url",
    # numbers
    "precio_mxn",
    "m2_construccion",
    "recamaras",
    "banos",
    "estacionamientos",
    "tipo",          # Loft / 1 rec / 2 rec ...
    "nuevo",         # Sí/No
    # text
    "fotos_urls",
    "pros",
    "contras",
    "notas",
    # flags
    "flag_tren_ruido",
    "flag_loft_sin_rec",
    "flag_fuera_presupuesto",
    "flag_pocos_deptos",
    # decision (app cliente)
    "decision_status",       # Pendiente / Apoya / Descarta / Visitar
    "decision_comentario",
    "decision_quien",
    "decision_fecha",
    # calculated
    "precio_por_m2",
]


DECISION_OPTIONS = ["Pendiente", "Apoya", "Descarta", "Visitar"]


def ensure_schema(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure all required columns exist and are in standard order."""
    df = df.copy()
    for col in DEFAULT_COLUMNS:
        if col not in df.columns:
            df[col] = None
    df = df[DEFAULT_COLUMNS]
    return df


def compute_fields(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # numeric coercions
    for c in ["precio_mxn", "m2_construccion", "recamaras", "banos", "estacionamientos"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # calculated
    df["precio_por_m2"] = df.apply(
        lambda r: (r["precio_mxn"] / r["m2_construccion"])
        if pd.notna(r["precio_mxn"]) and pd.notna(r["m2_construccion"]) and r["m2_construccion"] not in (0, 0.0)
        else None,
        axis=1,
    )

    # defaults
    if "decision_status" in df.columns:
        df["decision_status"] = df["decision_status"].fillna("Pendiente")
        df.loc[~df["decision_status"].isin(DECISION_OPTIONS), "decision_status"] = "Pendiente"

    return df


def read_deptos_excel(file) -> pd.DataFrame:
    """Read Deptos sheet -> DataFrame (schema enforced)."""
    df = pd.read_excel(file, sheet_name=SHEET_NAME, engine="openpyxl")
    df = ensure_schema(df)
    df = compute_fields(df)
    # id must exist
    if df["id"].isna().any():
        # rebuild missing ids safely
        df["id"] = range(1, len(df) + 1)
    df["id"] = df["id"].astype(int)
    return df


def _apply_formatting(ws):
    # Freeze header
    ws.freeze_panes = "A2"

    # header style
    header_fill = PatternFill("solid", fgColor="1F2937")  # dark
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # column widths (roughly)
    widths = {
        "A": 6,   # id
        "B": 34,  # nombre
        "C": 22,  # zona
        "D": 14,  # fuente
        "E": 42,  # url
        "F": 14,  # precio
        "G": 14,  # m2
        "H": 10,  # rec
        "I": 10,  # baños
        "J": 16,  # estac
        "K": 14,  # tipo
        "L": 10,  # nuevo
        "M": 42,  # fotos_urls
        "N": 34,  # pros
        "O": 34,  # contras
        "P": 28,  # notas
        "Q": 14,  # tren
        "R": 16,  # loft
        "S": 18,  # fuera pres
        "T": 16,  # pocos
        "U": 16,  # decision
        "V": 28,  # comentario
        "W": 18,  # quien
        "X": 18,  # fecha
        "Y": 14,  # $/m2
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Wrap text on big text cols
    wrap_cols = ["B", "M", "N", "O", "P", "V", "E"]
    for col in wrap_cols:
        for cell in ws[col]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Number formats
    money_cols = ["F"]
    ratio_cols = ["Y"]
    for col in money_cols:
        for cell in ws[col][1:]:
            cell.number_format = '"$"#,##0'
    for col in ratio_cols:
        for cell in ws[col][1:]:
            cell.number_format = '"$"#,##0.00'

    # Status coloring (column T = decision_status)
    status_fill = {
        "Pendiente": PatternFill("solid", fgColor="E5E7EB"),
        "Apoya": PatternFill("solid", fgColor="BBF7D0"),
        "Descarta": PatternFill("solid", fgColor="FECACA"),
        "Visitar": PatternFill("solid", fgColor="BFDBFE"),
    }
    # Find decision_status col index
    headers = [c.value for c in ws[1]]
    if "decision_status" in headers:
        idx = headers.index("decision_status") + 1
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=idx).value
            fill = status_fill.get(str(v), None)
            if fill:
                ws.cell(row=r, column=idx).fill = fill

    # Add an Excel table
    last_col = ws.max_column
    last_row = ws.max_row
    table_ref = f"A1:{_col_letter(last_col)}{last_row}"
    table = Table(displayName="DeptosTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def _col_letter(n: int) -> str:
    # 1-indexed
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)


def write_deptos_excel(df: pd.DataFrame, path_or_buffer) -> None:
    """Write Deptos sheet with formatting."""
    df = ensure_schema(df)
    df = compute_fields(df)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    _apply_formatting(ws)
    wb.save(path_or_buffer)


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")